#!/usr/bin/env python3
"""
季度销售数据 Excel 生成器
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建工作簿
wb = Workbook()
sheet = wb.active
sheet.title = "季度销售数据"

# 定义样式
header_font = Font(name='Arial', bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2C3E50")
header_alignment = Alignment(horizontal='center', vertical='center')

total_font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
total_fill = PatternFill("solid", fgColor="27AE60")
total_alignment = Alignment(horizontal='center', vertical='center')

data_alignment = Alignment(horizontal='right', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置标题
sheet['A1'] = "2024年度季度销售报告"
sheet.merge_cells('A1:E1')
sheet['A1'].font = Font(name='Arial', bold=True, size=16, color="2C3E50")
sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

# 表头
headers = ['季度', '产品A', '产品B', '产品C', '季度合计']
for col_idx, header in enumerate(headers, start=1):
    cell = sheet.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# 数据行
data = [
    ['Q1', 125000, 89000, 156000],
    ['Q2', 134000, 92000, 168000],
    ['Q3', 148000, 98000, 175000],
    ['Q4', 162000, 105000, 189000]
]

for row_idx, row_data in enumerate(data, start=4):
    sheet.cell(row=row_idx, column=1, value=row_data[0]).alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(2, 5):
        cell = sheet.cell(row=row_idx, column=col_idx, value=row_data[col_idx-1])
        cell.alignment = data_alignment
        cell.border = border
        cell.number_format = '#,##0'
    
    # 季度合计公式
    total_cell = sheet.cell(row=row_idx, column=5)
    total_cell.value = f'=SUM(B{row_idx}:D{row_idx})'
    total_cell.alignment = data_alignment
    total_cell.border = border
    total_cell.number_format = '#,##0'

# 合计行
total_row = 8
sheet.cell(row=total_row, column=1, value='年度合计').font = total_font
sheet.cell(row=total_row, column=1).fill = total_fill
sheet.cell(row=total_row, column=1).alignment = total_alignment
sheet.cell(row=total_row, column=1).border = border

for col_idx in range(2, 6):
    cell = sheet.cell(row=total_row, column=col_idx)
    cell.value = f'=SUM({chr(64+col_idx)}4:{chr(64+col_idx)}7)'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    cell.number_format = '#,##0'

# 列宽设置
sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15

# 行高设置
sheet.row_dimensions[1].height = 30
for row_idx in range(3, 9):
    sheet.row_dimensions[row_idx].height = 25

# 保存文件
output_file = "/tmp/quarterly_sales_report.xlsx"
wb.save(output_file)
print(f"Excel文件已创建: {output_file}")